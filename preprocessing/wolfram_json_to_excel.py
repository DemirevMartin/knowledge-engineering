"""
Wolfram Food Carbon Footprint — JSON to Excel Converter
========================================================
Converts the raw Wolfram Language serialized JSON file to a clean,
formatted Excel file with ingredient names and CO2e values.

Usage:
    python wolfram_json_to_excel.py

Input:  Food-Carbon-Footprint.json   (Wolfram serialized expression)
Output: Wolfram_Food_Carbon_Footprint.xlsx

Output columns: #, Clean Name, CO2e Min (kg), CO2e Max (kg), CO2e Avg (kg), CO2e Range
"""

import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
INPUT_FILE  = "Food-Carbon-Footprint.json"
OUTPUT_FILE = "Wolfram_Food_Carbon_Footprint_trial.xlsx"

GREEN   = "1E7B4B"
WHITE   = "FFFFFF"
LTGREEN = "E8F5EE"


# ─────────────────────────────────────────────
# STEP 1: Read file at byte level
# ─────────────────────────────────────────────
# MUST read as bytes — names are stored as a packed binary string
# where byte positions (not character positions) define name boundaries.
with open(INPUT_FILE, "rb") as f:
    raw_bytes = f.read()

# Decode as latin-1 for structural regex parsing (safe for all byte values)
raw_text = raw_bytes.decode("latin-1")


# ─────────────────────────────────────────────
# STEP 2: Extract ingredient names
# ─────────────────────────────────────────────

# 2a. Extract the 539 byte offsets from the Wolfram List structure:
#     ["List", 2, ["List", 0, 4, 19, 24, ...], "'beefdark chocolates...'"]
offsets_match = re.search(r'\["List",2,\["List",([\d,\s]+)\]', raw_text)
offsets = [int(x.strip()) for x in offsets_match.group(1).split(",")]

# 2b. Locate the packed name byte string — all 538 names concatenated with NO separators
name_start = raw_bytes.find(b"'beef") + 1        # +1 to skip opening quote
name_end   = raw_bytes.find(b"'\"", name_start)  # closing quote sequence
name_bytes = raw_bytes[name_start:name_end]

# 2c. Slice each name using its byte offset pair, decode as UTF-8
names = [
    name_bytes[offsets[i]:offsets[i + 1]].decode("utf-8", errors="replace")
    for i in range(len(offsets) - 1)
]

# 2d. Normalize: title-case (Wolfram uses lowercase plurals e.g. "dark chocolates")
names = [n.title() for n in names]


# ─────────────────────────────────────────────
# STEP 3: Extract CO2e values
# ─────────────────────────────────────────────

# 3a. CO2e values stored as Wolfram Interval objects: ["Interval", ["List", min, max]]
all_intervals = [
    (float(m.group(1)), float(m.group(2)))
    for m in re.finditer(
        r'\["Interval",\["List",([\d.e+-]+),([\d.e+-]+)\]\]',
        raw_text
    )
]

# 3b. File contains 1,076 intervals (538 x 2) — Wolfram stores data twice.
#     Keep only the first 538 (actual data); discard the cached duplicate.
intervals = all_intervals[:538]


# ─────────────────────────────────────────────
# STEP 4: Align, derive columns, sort
# ─────────────────────────────────────────────
assert len(names) == len(intervals) == 538, \
    f"Count mismatch: {len(names)} names vs {len(intervals)} intervals"

df = pd.DataFrame({
    "Clean Name":    names,
    "CO2e Min (kg)": [mn for mn, _ in intervals],
    "CO2e Max (kg)": [mx for _, mx in intervals],
})
df["CO2e Avg (kg)"] = ((df["CO2e Min (kg)"] + df["CO2e Max (kg)"]) / 2).round(2)
df["CO2e Range"]    = df["CO2e Min (kg)"].round(2).astype(str) + " - " + df["CO2e Max (kg)"].round(2).astype(str)

# Sort alphabetically A → Z
df = df.sort_values("Clean Name").reset_index(drop=True)
df.index += 1  # 1-based row numbers for the # column


# ─────────────────────────────────────────────
# STEP 5: Export to Excel
# ─────────────────────────────────────────────

# 5a. Bulk write with pandas — columns: #, Clean Name, CO2e Min, CO2e Max, CO2e Avg, CO2e Range
df.to_excel(OUTPUT_FILE, index=True, index_label="#", sheet_name="Carbon Footprint Data")

# 5b. Format with openpyxl
wb = load_workbook(OUTPUT_FILE)
ws = wb.active

# Style header row (row 1)
for cell in ws[1]:
    cell.font      = Font(name="Arial", bold=True, color=WHITE, size=11)
    cell.fill      = PatternFill("solid", start_color=GREEN)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 28

# Style data rows with CO2e-level color coding
thin   = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    avg_val = row[4].value  # column E = CO2e Avg (kg)
    if isinstance(avg_val, (int, float)):
        if avg_val >= 20:   bg = "FFD6D6"   # red    — very high
        elif avg_val >= 10: bg = "FFE8CC"   # orange — high
        elif avg_val >= 5:  bg = "FFFACC"   # yellow — medium
        else:               bg = "FFFFFF" if row[0].row % 2 == 0 else LTGREEN
    else:
        bg = "FFFFFF"

    for cell in row:
        cell.fill      = PatternFill("solid", start_color=bg)
        cell.font      = Font(name="Arial", size=10)
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row[1].alignment = Alignment(horizontal="left", vertical="center")  # Clean Name left-aligned

# Column widths: #, Clean Name, CO2e Min, CO2e Max, CO2e Avg, CO2e Range
for i, w in enumerate([5, 28, 16, 16, 16, 18], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Legend at bottom
lr = ws.max_row + 2
ws.cell(row=lr, column=1, value="Legend:").font = Font(name="Arial", bold=True, size=10)
for i, (color, label) in enumerate([
    ("FFD6D6", "Very High CO2e (>= 20 kg/kg)"),
    ("FFE8CC", "High CO2e (10-19 kg/kg)"),
    ("FFFACC", "Medium CO2e (5-9 kg/kg)"),
    (LTGREEN,  "Low CO2e (< 5 kg/kg)"),
], 1):
    ws.cell(row=lr + i, column=1).fill   = PatternFill("solid", start_color=color)
    ws.cell(row=lr + i, column=1).border = border
    ws.cell(row=lr + i, column=2, value=label).font = Font(name="Arial", size=10)

ws.freeze_panes = "A2"
wb.save(OUTPUT_FILE)

print(f"Done! {len(df)} rows written to {OUTPUT_FILE}")
print(f"Columns: {['#'] + df.columns.tolist()}")
